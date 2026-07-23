import os
from openpyxl import load_workbook

EXCEL_FILE = "clan_2527.xlsx"


def backfill():
    if not os.path.exists(EXCEL_FILE):
        print(f"{EXCEL_FILE} not found")
        return

    wb = load_workbook(EXCEL_FILE)
    sheets = [s for s in wb.worksheets if s.title != "Sheet1"]
    sheet_names = sorted([s.title for s in sheets])

    updated = 0
    for i in range(1, len(sheet_names)):
        prev_name = sheet_names[i - 1]
        curr_name = sheet_names[i]
        curr_ws = wb[curr_name]

        if curr_ws["D2"].value is not None:
            continue

        prev_reps = {}
        for row in wb[prev_name].iter_rows(min_row=4, max_col=2, values_only=True):
            name, rep = row[0], row[1]
            if name is not None and rep is not None:
                prev_reps[str(name).strip()] = int(float(rep))

        curr_reps = {}
        for row in curr_ws.iter_rows(min_row=4, max_col=2, values_only=True):
            name, rep = row[0], row[1]
            if name is not None and rep is not None:
                curr_reps[str(name).strip()] = int(float(rep))

        day_points = 0
        for name, curr_rep in curr_reps.items():
            if name in prev_reps:
                diff = curr_rep - prev_reps[name]
                if diff > 0:
                    day_points += diff

        curr_ws["D1"] = "Day Points"
        curr_ws["D2"] = day_points
        updated += 1
        prev_date = prev_name
        print(f"  {curr_name}: {day_points:>6}  (vs {prev_date})")

    if updated:
        wb.save(EXCEL_FILE)
        print(f"\nBackfilled {updated} sheets")
    else:
        print("No sheets needed updating")


if __name__ == "__main__":
    backfill()
