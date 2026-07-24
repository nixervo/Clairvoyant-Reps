import os
import sys
from openpyxl import load_workbook

EXCEL_FILE = "clan_2527.xlsx"
SEASON_BOUNDARY = "2026-07-19"  # Jul 19 and before = S61, after = S62


def fix_xlsx_tags():
    wb = load_workbook(EXCEL_FILE)
    sheets = [s for s in wb.worksheets if s.title != "Sheet1"]

    fixed = 0
    for ws in sheets:
        a1 = ws["A1"].value or ""
        title = ws.title

        current_season = None
        if "|" in a1:
            meta = a1.rsplit("|", 1)[1].strip()
            if meta.startswith("S"):
                current_season = int(meta[1:])

        correct_season = 61 if title <= SEASON_BOUNDARY else 62

        if current_season != correct_season:
            base = a1.split("|")[0].strip() if "|" in a1 else a1
            ws["A1"] = f"{base} | S{correct_season}"
            fixed += 1
            print(f"  {title}: S{current_season} -> S{correct_season}")

    wb.save(EXCEL_FILE)
    print(f"\nFixed {fixed} sheets")
    return fixed


def delete_history_files():
    deleted = 0
    for f in os.listdir("."):
        if f.startswith("history_") and f.endswith(".html"):
            os.remove(f)
            deleted += 1
        if f == "history.html":
            os.remove(f)
            deleted += 1
    print(f"Deleted {deleted} history files")


def regenerate_history():
    sys.path.insert(0, ".")
    from clan_snapshot import save_daily_history
    save_daily_history()


if __name__ == "__main__":
    print("=== Fixing season tags ===")
    fix_xlsx_tags()
    print("\n=== Deleting old history files ===")
    delete_history_files()
    print("\n=== Regenerating history pages ===")
    regenerate_history()
    print("\nDone!")
