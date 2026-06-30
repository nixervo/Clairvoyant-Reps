import os
import urllib.request
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import json
import time
import sys
import base64

API_URL = "https://playninjarift.com/api/detail_clan_website.php?clan_id=2527"
TARGET_TZ = timezone(timedelta(hours=8))
EXCEL_FILE = "clan_2527.xlsx"
HOURLY_CACHE = "_hourly_cache.json"
CACHE_30M = "_30m_cache.json"
CLAN_ID = 2527

def fetch_clan():
    req = urllib.request.Request(API_URL, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode())

SEASON_API = "https://playninjarift.com/api/refresh_time_website.php"
RANKING_API = "https://playninjarift.com/api/clan_ranking_website.php"

def fetch_season_info():
    req = urllib.request.Request(SEASON_API, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode())

def fetch_clan_ranking():
    req = urllib.request.Request(RANKING_API, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode())
    for entry in data:
        if entry["clan_id"] == CLAN_ID:
            return entry
    return {}

def get_previous_sheet_names(wb):
    names = [s.title for s in wb.worksheets]
    names.sort()
    return names

def load_prev_from_xlsx(filename, before_date):
    prev_data = []
    prev_timestamp = None
    if not os.path.exists(filename):
        return prev_data, prev_timestamp
    wb = load_workbook(filename)
    prev_names = [s.title for s in wb.worksheets]
    prev_names.sort()
    prev_sheet_name = None
    for n in reversed(prev_names):
        if n < before_date:
            prev_sheet_name = n
            break
    if prev_sheet_name and prev_sheet_name in wb.sheetnames:
        ps = wb[prev_sheet_name]
        raw = ps["A2"].value
        if raw:
            prev_timestamp = raw.replace("Timestamp: ", "")
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[0] and row[1] is not None:
                prev_data.append({"character_name": row[0], "member_reputation": int(row[1])})
    return prev_data, prev_timestamp

def load_hourly_cache():
    if not os.path.exists(HOURLY_CACHE):
        return {}, None
    with open(HOURLY_CACHE, encoding="utf-8") as f:
        c = json.load(f)
    return c.get("members", {}), c.get("timestamp")

def save_hourly_cache(members, now):
    cache = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "members": {m["character_name"]: m["member_reputation"] for m in members}
    }
    with open(HOURLY_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_30m_cache():
    if not os.path.exists(CACHE_30M):
        return None
    with open(CACHE_30M, encoding="utf-8") as f:
        return json.load(f)

def save_30m_cache(members, now):
    cache = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "members": {m["character_name"]: m["member_reputation"] for m in members}
    }
    with open(CACHE_30M, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def compute_rolling_avg_daily_gain(filename, before_date):
    if not os.path.exists(filename):
        return None
    wb = load_workbook(filename)
    names = sorted([s.title for s in wb.worksheets if s.title != "Sheet1" and s.title < before_date])
    gains = []
    for i in range(1, len(names)):
        prev_total = 0
        curr_total = 0
        ps = wb[names[i-1]]
        cs = wb[names[i]]
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                prev_total += int(row[1])
        for row in cs.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                curr_total += int(row[1])
        gains.append(curr_total - prev_total)
    if not gains:
        return None
    return sum(gains) / len(gains)

def compute_changes(members, prev_data):
    prev_names = {m["character_name"] for m in prev_data}
    today_names = {m["character_name"] for m in members}
    left_names = sorted(prev_names - today_names)
    joined_names = sorted(today_names - prev_names)
    return left_names, joined_names

def compute_diff(members, prev_data):
    prev_map = {m["character_name"]: m["member_reputation"] for m in prev_data}
    result = []
    for m in members:
        name = m["character_name"]
        reps = m["member_reputation"]
        if name in prev_map:
            diff = reps - prev_map[name]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
        else:
            diff_str = "N/A"
        result.append((name, reps, diff_str))
    return result

def write_sheet(ws, data, prev_data, now):
    rows = compute_diff(data["members"], prev_data)
    ws.title = now.strftime("%Y-%m-%d")

    names = [r[0] for r in rows]
    reps = [str(r[1]) for r in rows]
    diffs = [r[2] for r in rows]

    max_name = max((len(n) for n in names), default=10)
    max_reps = max((len(r) for r in reps), default=4)
    max_diff = max((len(d) for d in diffs), default=3)

    ws.column_dimensions["A"].width = max(max_name + 5, 10)
    ws.column_dimensions["B"].width = max(max_reps + 3, 8)
    ws.column_dimensions["C"].width = max(max_diff + 3, 14)

    clan_name = data.get("clan_name", "Unknown")
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Clan: {clan_name} ({CLAN_ID})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Timestamp: {now.strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Name", "Total Reps", "Daily Reps (+1d)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (name, reps_val, diff_val) in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=2, value=reps_val).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=diff_val).alignment = Alignment(horizontal="center", vertical="center")

def save_xlsx(data, prev_data, now):
    sheet_name = now.strftime("%Y-%m-%d")

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, data, prev_data, now)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(EXCEL_FILE)
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Saved sheet '{sheet_name}' to {EXCEL_FILE}")

def save_seasonal_xlsx(members, season_num):
    filename = f"S{season_num}_ID{CLAN_ID}.xlsx"
    if os.path.exists(filename):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = f"Season {season_num}"
    header = f"[S{season_num}] Total Reps"
    ws["A1"] = "Name"
    ws["B1"] = header
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    max_name = 10
    max_reps = 4
    for idx, m in enumerate(members, 2):
        name = m["character_name"]
        reps = m["member_reputation"]
        ws.cell(row=idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=idx, column=2, value=reps).alignment = Alignment(horizontal="center", vertical="center")
        max_name = max(max_name, len(name))
        max_reps = max(max_reps, len(str(reps)))
    ws.column_dimensions["A"].width = max_name + 5
    ws.column_dimensions["B"].width = max_reps + 3
    wb.save(filename)
    print(f"Saved seasonal snapshot: {filename}")

def compute_season_projection(clan_reputation, avg_daily_gain, season_end_dt, now):
    if avg_daily_gain is None or avg_daily_gain <= 0:
        return None
    projection = clan_reputation
    total_days = 0
    current = now + timedelta(days=1)
    current = current.replace(hour=0, minute=0, second=0, microsecond=0)
    while current <= season_end_dt:
        if current.weekday() >= 5:
            projection += avg_daily_gain * 2
        else:
            projection += avg_daily_gain
        total_days += 1
        current += timedelta(days=1)
    days_left = (season_end_dt - now).days
    return {"projection": int(round(projection)), "avg_daily": int(round(avg_daily_gain)), "days_left": days_left}

def diff_html(diff_str):
    if diff_str.startswith("+"):
        return f'<span class="up">{diff_str}</span>'
    elif diff_str.startswith("-"):
        return f'<span class="down">{diff_str}</span>'
    else:
        return f'<span class="na">{diff_str}</span>'

def save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, all_dates, show_changes, season_info=None, stats=None, diff_30m=None):
    daily_rows = compute_diff(data["members"], prev_data)
    clan_name = data.get("clan_name", "Unknown")
    date_str = now.strftime("%Y-%m-%d")
    ts_str = now.strftime("%Y-%m-%d %H:%M:%S")
    member_count = len(data["members"])

    logo_b64 = ""
    logo_path = "clan_logo.png"
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()

    favicon_b64 = ""
    if os.path.exists("favicon.ico"):
        with open("favicon.ico", "rb") as f:
            favicon_b64 = base64.b64encode(f.read()).decode()

    archive_links = "".join(
        f'<a href="{d}.html" class="{"active" if d == date_str else ""}">{d}</a>'
        for d in sorted(all_dates, reverse=True)
    )

    diffs_30m_map = diff_30m.get("diffs", {}) if isinstance(diff_30m, dict) else {}
    table_rows = "".join(
        f"<tr><td>{name}</td><td class=\"num\">{reps}</td><td class=\"num\">{diff_html(diffs_30m_map.get(name, 'N/A'))}</td><td class=\"num\">{diff_html(hourly_diffs.get(name, 'N/A'))}</td><td class=\"num\">{diff_html(daily_diff)}</td></tr>"
        for name, reps, daily_diff in daily_rows
    )

    changes_html = ""
    if show_changes and prev_data:
        left_names, joined_names = compute_changes(data["members"], prev_data)
        if left_names or joined_names:
            left_items = "".join(f"<li>{n}</li>" for n in left_names)
            joined_items = "".join(f"<li>{n}</li>" for n in joined_names)
            changes_html = f"""
  <div class="changes">
    <div class="changes-title">Member Changes (from {prev_timestamp or 'previous snapshot'})</div>
    <div class="changes-cols">
      <div class="changes-col">
        <div class="changes-head left">Left ({len(left_names)})</div>
        <ul>{left_items}</ul>
      </div>
      <div class="changes-col">
        <div class="changes-head joined">Joined ({len(joined_names)})</div>
        <ul>{joined_items}</ul>
      </div>
    </div>
  </div>"""

    logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo" alt="Clairvoyant">' if logo_b64 else ""
    favicon_html = f'<link rel="icon" type="image/x-icon" href="data:image/x-icon;base64,{favicon_b64}">' if favicon_b64 else ""

    hourly_ref = f"Ref (hourly): {hourly_ts}" if hourly_ts else ""
    ref_30m_val = diff_30m.get("ts", "") if isinstance(diff_30m, dict) else ""
    ref_30m = f"Ref (30m): {ref_30m_val}" if ref_30m_val else ""
    daily_ref = f"Ref (daily): {prev_timestamp}" if prev_timestamp else ""

    timer_html = ""
    season_end_iso = ""
    if season_info:
        season_num = season_info["season"]
        end_dt = datetime.strptime(season_info["season_end"], "%Y-%m-%d %H:%M:%S").replace(tzinfo=TARGET_TZ)
        season_end_iso = end_dt.strftime("%Y-%m-%dT%H:%M:%S%z")
        timer_html = f"""
  <div class="timer-bar">
    <span class="timer-season">Season <span id="season-num">{season_num}</span></span>
    <span class="timer-sep">&middot;</span>
    <span class="timer-clock">
      <span class="timer-digits"><span id="timer-d">--</span><span class="timer-unit">d</span></span>
      <span class="timer-digits"><span id="timer-h">--</span><span class="timer-unit">h</span></span>
      <span class="timer-digits"><span id="timer-m">--</span><span class="timer-unit">m</span></span>
      <span class="timer-digits"><span id="timer-s">--</span><span class="timer-unit">s</span></span>
    </span>
  </div>"""

    stats_html = ""
    if stats:
        stats_html = f"""
  <div class="stats-bar">
    <div class="stats-col">
      <span class="stat-label">Today</span>
      <span class="stat-val" id="today-gain">+{stats['today_gain']:,}</span>
      <span class="stat-label">Season Total</span>
      <span class="stat-val">{stats['season_total']:,}</span>
    </div>
    <div class="stats-col">
      <span class="stat-label">Est. Season Total</span>
      <span class="stat-val">{stats['projection']:,}</span>
      <span class="stat-label">Avg/Day &middot; {stats['days_left']}d left</span>
      <span class="stat-val">{stats['avg_daily']:,}</span>
    </div>
  </div>"""

    script_html = ""
    if season_info:
        script_html = """<script>
(function() {
  var end = new Date(\"""" + season_end_iso + """\").getTime();
  function tick() {
    var diff = end - new Date().getTime();
    if (diff <= 0) { document.getElementById("timer-d").textContent = "0"; document.getElementById("timer-h").textContent = "00"; document.getElementById("timer-m").textContent = "00"; document.getElementById("timer-s").textContent = "00"; return; }
    document.getElementById("timer-d").textContent = Math.floor(diff / 86400000);
    document.getElementById("timer-h").textContent = String(Math.floor((diff % 86400000) / 3600000)).padStart(2,"0");
    document.getElementById("timer-m").textContent = String(Math.floor((diff % 3600000) / 60000)).padStart(2,"0");
    document.getElementById("timer-s").textContent = String(Math.floor((diff % 60000) / 1000)).padStart(2,"0");
  }
  tick();
  setInterval(tick, 1000);
})();
(function() {
  var tbody = document.querySelector("tbody");
  var defaultRows = tbody.innerHTML;
  var sortCol = -1, sortDir = 0;
  var ths = document.querySelectorAll("th");
  for (var i = 0; i < ths.length; i++) (function(col) {
    ths[col].addEventListener("click", function() {
      if (sortCol !== col) { sortCol = col; sortDir = 1; }
      else { sortDir = (sortDir + 1) % 3; if (sortDir === 0) { tbody.innerHTML = defaultRows; for (var a = 0; a < ths.length; a++) ths[a].querySelector(".sort-arrow").textContent = ""; return; } }
      for (var a = 0; a < ths.length; a++) ths[a].querySelector(".sort-arrow").textContent = "";
      ths[col].querySelector(".sort-arrow").textContent = sortDir === 1 ? "\\u25B2" : "\\u25BC";
      var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
      rows.sort(function(a, b) {
        var va = a.cells[col].textContent.trim(), vb = b.cells[col].textContent.trim();
        if (col === 0) return sortDir === 1 ? va.localeCompare(vb) : vb.localeCompare(va);
        var na = parseFloat(va) || -1/0, nb = parseFloat(vb) || -1/0;
        return sortDir === 1 ? na - nb : nb - na;
      });
      for (var r = 0; r < rows.length; r++) tbody.appendChild(rows[r]);
    });
  })(i);
})();
</script>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{favicon_html}
<title>{clan_name} [Reps]</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    background: #080810;
    color: #e0e0e0;
    min-height: 100vh;
    display: flex;
    justify-content: center;
    padding: 32px 16px;
  }}
  .container {{
    max-width: 960px;
    width: 100%;
    box-shadow: 0 0 40px rgba(233, 69, 96, 0.06), 0 8px 32px rgba(0,0,0,0.5);
    border-radius: 16px;
    overflow: hidden;
  }}
  .header {{
    text-align: center;
    padding: 32px 24px 24px;
    background: linear-gradient(135deg, #0f0f1e 0%, #1a1a30 50%, #0d1b2a 100%);
    position: relative;
    overflow: hidden;
  }}
  .header::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #e94560, #ff6b8a, #e94560);
    background-size: 200% 100%;
    animation: shimmer 3s ease-in-out infinite;
  }}
  @keyframes shimmer {{ 0%,100% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} }}
  .logo {{
    width: 132px; height: 132px;
    object-fit: contain;
    margin-bottom: 16px;
    filter: drop-shadow(0 0 16px rgba(233, 69, 96, 0.3));
  }}
  .header h1 {{
    font-size: 30px;
    font-weight: 700;
    color: #fff;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
  }}
  .header .sub {{
    font-size: 17px;
    color: #888;
    display: flex;
    justify-content: center;
    gap: 16px;
    flex-wrap: wrap;
  }}
  .header .sub span {{ color: #aaa; }}
  .archive {{
    display: flex;
    gap: 6px;
    justify-content: center;
    padding: 14px 20px;
    background: #0c0c18;
    border-bottom: 1px solid #1a1a2e;
    border-top: 1px solid #1a1a2e;
  }}
  .archive a {{
    color: #777;
    text-decoration: none;
    font-size: 12px;
    padding: 5px 14px;
    border-radius: 20px;
    border: 1px solid #1a1a2e;
    transition: 0.25s;
  }}
  .archive a:hover {{ border-color: #e94560; color: #fff; background: rgba(233, 69, 96, 0.08); }}
  .archive a.active {{ border-color: #e94560; color: #fff; background: #e94560; font-weight: 600; }}
  .table-wrap {{ overflow-x: auto; }}
  table {{
    width: 100%;
    min-width: 0;
    border-collapse: collapse;
    background: #0c0c14;
  }}
  thead {{ position: sticky; top: 0; z-index: 1; }}
  th {{
    background: #0f0f1e;
    padding: 14px 18px;
    text-align: center;
    font-size: 15px;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #e94560;
    font-weight: 600;
  }}
  td {{
    padding: 11px 18px;
    border-bottom: 1px solid #14141f;
    font-size: 14px;
    color: #ccc;
    text-align: center;
  }}
  tr:nth-child(even) td {{ background: rgba(255,255,255,0.015); }}
  tr:hover td {{ background: rgba(233, 69, 96, 0.04); }}
  td.num {{ font-variant-numeric: tabular-nums; }}
  .up {{ color: #4caf50; }} .up::before {{ content: '\\25B2 '; font-size: 10px; }}
  .down {{ color: #f44336; }} .down::before {{ content: '\\25BC '; font-size: 10px; }}
  .na {{ color: #555; }}
  .changes {{
    background: #0c0c14;
    padding: 20px 24px;
    border-top: 1px solid #14141f;
  }}
  .changes-title {{
    font-size: 13px;
    color: #e94560;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 14px;
    text-align: center;
  }}
  .changes-cols {{
    display: flex;
    gap: 24px;
    justify-content: center;
  }}
  .changes-col {{
    flex: 1;
    max-width: 320px;
  }}
  .changes-head {{
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-weight: 600;
    margin-bottom: 8px;
    text-align: center;
  }}
  .changes-head.left {{ color: #f44336; }}
  .changes-head.joined {{ color: #4caf50; }}
  .changes-col ul {{
    list-style: none;
    padding: 0;
    margin: 0;
    text-align: center;
  }}
  .changes-col li {{
    padding: 4px 0;
    font-size: 14px;
    color: #ccc;
    border-bottom: 1px solid #14141f;
  }}
  .changes-col li:last-child {{ border-bottom: none; }}
  .footer {{
    text-align: center;
    padding: 18px 20px;
    background: #08080f;
    color: #444;
    font-size: 12px;
    border-top: 1px solid #12121e;
  }}
  .footer .ref {{ color: #555; font-size: 11px; margin-top: 2px; }}
  .footer a {{ color: #e94560; text-decoration: none; }}
  .timer-bar {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    padding: 12px 20px;
    background: #0f142373;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
    font-size: 15px;
    flex-wrap: wrap;
  }}
  .timer-season {{ color: #eab308; font-weight: 700; letter-spacing: 0.5px; }}
  .timer-sep {{ color: #444; }}
  .timer-clock {{ display: flex; align-items: center; gap: 6px; }}
  .timer-digits {{ font-variant-numeric: tabular-nums; }}
  .timer-digits span:first-child {{ color: #2dd4bf; font-weight: 600; min-width: 28px; display: inline-block; text-align: center; }}
  .timer-unit {{ color: #888; font-size: 12px; margin-left: 1px; }}
  .stats-bar {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 32px;
    padding: 14px 20px;
    background: #0f142373;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
    flex-wrap: wrap;
  }}
  .stats-col {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 2px;
  }}
  .stat-label {{ color: #888; font-size: 12px; text-transform: uppercase; letter-spacing: 0.3px; }}
  .stat-val {{ color: #e0e0e0; font-size: 18px; font-weight: 700; font-variant-numeric: tabular-nums; }}
  #today-gain {{ color: #4caf50; }}
  th {{ cursor: pointer; user-select: none; }}
  th .sort-arrow {{ font-size: 11px; margin-left: 3px; }}
  @media (max-width: 600px) {{
    body {{ padding: 16px 8px; }}
    .header {{ padding: 24px 16px 20px; }}
    .header h1 {{ font-size: 22px; }}
    .header .sub {{ font-size: 14px; }}
    .logo {{ width: 96px; height: 96px; }}
    table {{ min-width: 520px; }}
    th, td {{ padding: 10px 10px; font-size: 12px; }}
    .changes {{ padding: 16px; }}
    .changes-cols {{ flex-direction: column; gap: 14px; }}
    .changes-col {{ max-width: 100%; }}
    .archive {{ flex-wrap: nowrap; overflow-x: auto; justify-content: flex-start; -webkit-overflow-scrolling: touch; scrollbar-width: none; }}
    .archive::-webkit-scrollbar {{ display: none; }}
    .archive a {{ flex-shrink: 0; }}
    .stats-bar {{ gap: 16px; padding: 12px 16px; }}
    .stat-val {{ font-size: 15px; }}
    .stats-col {{ width: 100%; }}
    .stats-col + .stats-col {{ border-top: 1px solid #1a1a2e; padding-top: 10px; }}
  }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    {logo_html}
    <h1>{clan_name}</h1>
    <div class="sub">
      <span>Clan ID: {CLAN_ID}</span>
      <span>&middot;</span>
      <span>{member_count} members</span>
    </div>
  </div>
  {timer_html}
  {stats_html}
  {f'<div class="archive">{archive_links}</div>' if archive_links else ""}
  <div class="table-wrap">
  <table>
    <thead><tr><th>Name <span class="sort-arrow"></span></th><th>Total Reps <span class="sort-arrow"></span></th><th>30m Reps <span class="sort-arrow"></span></th><th>Hourly Reps (+1h) <span class="sort-arrow"></span></th><th>Daily Reps (+1d) <span class="sort-arrow"></span></th></tr></thead>
    <tbody>{table_rows}</tbody>
  </table>
  </div>
  {changes_html}
  <div class="footer">
    Snapshot: {ts_str}
    <div class="ref">{ref_30m}{" &middot; " if ref_30m and (hourly_ref or daily_ref) else ""}{hourly_ref}{" &middot; " if hourly_ref and daily_ref else ""}{daily_ref}</div>
  </div>
</div>
{script_html}
</body>
</html>"""

    index_path = "index.html"
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[{ts_str}] Saved {index_path}")

    if all_dates:
        archive_path = f"{date_str}.html"
        with open(archive_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"[{ts_str}] Saved {archive_path}")

def save_snapshot(data):
    now = datetime.now(TARGET_TZ)
    is_daily = (now.hour == 13)
    sheet_name = now.strftime("%Y-%m-%d")

    prev_data, prev_timestamp = load_prev_from_xlsx(EXCEL_FILE, sheet_name)

    hourly_cache, hourly_ts = load_hourly_cache()
    hourly_diffs = {}
    for m in data["members"]:
        name = m["character_name"]
        reps = m["member_reputation"]
        if name in hourly_cache:
            diff = reps - hourly_cache[name]
            hourly_diffs[name] = f"+{diff}" if diff > 0 else str(diff)
        else:
            hourly_diffs[name] = "N/A"

    cache_30m = load_30m_cache()
    diffs_30m = {}
    ref_30m_ts = ""
    if cache_30m:
        ref_30m_ts = cache_30m.get("timestamp", "")
        for m in data["members"]:
            name = m["character_name"]
            reps = m["member_reputation"]
            if name in cache_30m.get("members", {}):
                diff = reps - cache_30m["members"][name]
                diffs_30m[name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                diffs_30m[name] = "N/A"
    diff_30m_data = {"ts": ref_30m_ts, "diffs": diffs_30m}

    try:
        season_info = fetch_season_info()
    except Exception:
        season_info = None

    try:
        ranking = fetch_clan_ranking()
        clan_reputation = ranking.get("clan_reputation", 0)
        today_gain = ranking.get("clan_day_points", 0)
    except Exception:
        clan_reputation = 0
        today_gain = 0

    stats = None
    if season_info:
        end_dt = datetime.strptime(season_info["season_end"], "%Y-%m-%d %H:%M:%S").replace(tzinfo=TARGET_TZ)
        avg_daily = compute_rolling_avg_daily_gain(EXCEL_FILE, sheet_name)
        proj = compute_season_projection(clan_reputation, avg_daily, end_dt, now)
        if proj:
            stats = {"today_gain": today_gain, "season_total": clan_reputation, "projection": proj["projection"], "avg_daily": proj["avg_daily"], "days_left": proj["days_left"]}

    is_hourly_mark = (now.minute <= 1)

    if is_daily:
        save_xlsx(data, prev_data, now)
        existing_html = [f.replace(".html", "") for f in os.listdir(".") if f.endswith(".html") and f[:4].isdigit() and f != "index.html"]
        all_dates = set(existing_html)
        all_dates.add(sheet_name)
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, sorted(all_dates), show_changes=True, season_info=season_info, stats=stats, diff_30m=diff_30m_data)
    else:
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, [], show_changes=False, season_info=season_info, stats=stats, diff_30m=diff_30m_data)

    save_30m_cache(data["members"], now)
    if is_hourly_mark:
        save_hourly_cache(data["members"], now)

    if season_info and now >= end_dt and avg_daily is not None:
        save_seasonal_xlsx(data["members"], season_info["season"])

def run():
    print("Clan snapshot daemon started. Running every 30 minutes.")
    while True:
        now = datetime.now(TARGET_TZ)
        if now.minute < 1:
            target = now.replace(minute=1, second=0, microsecond=0)
        elif now.minute < 31:
            target = now.replace(minute=31, second=0, microsecond=0)
        else:
            target = (now + timedelta(hours=1)).replace(minute=1, second=0, microsecond=0)
        sleep_sec = (target - now).total_seconds()
        time.sleep(sleep_sec)
        try:
            data = fetch_clan()
            save_snapshot(data)
            ts = datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S")
            os.system("git add -A")
            os.system(f'git commit -m "auto: snapshot {ts}" --allow-empty')
            os.system("git push")
        except Exception as e:
            print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] ERROR: {e}")

def fetch_once():
    data = fetch_clan()
    save_snapshot(data)

if __name__ == "__main__":
    if "--once" in sys.argv:
        fetch_once()
    else:
        run()
